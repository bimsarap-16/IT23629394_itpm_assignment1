from playwright.sync_api import sync_playwright
import os
import re
import openpyxl
import argparse

def run_test(excel_file, url):
    if not os.path.exists(excel_file):
        print(f"File not found: {excel_file}")
        return

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["Test cases"] if "Test cases" in wb.sheetnames else wb.active

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        page.goto(url, wait_until="networkidle")
        input_box = page.locator("textarea").first
        
        for i in range(2, ws.max_row + 1):
            input_text = ws.cell(i, 3).value 
            if not input_text: continue

            text_to_type = str(input_text).strip()
            
            try:
                clear_btn = page.get_by_role("button", name=re.compile("Clear", re.I))
                if clear_btn.is_visible():
                    clear_btn.click()
                else:
                    input_box.fill("")
                
                input_box.click()
                page.keyboard.type(text_to_type, delay=80) 
                page.keyboard.press("Space")
                
                page.wait_for_timeout(3000) 
                    
                page_text = page.inner_text("body")
                lines = [line.strip() for line in page_text.split('\n') if line.strip()]
                
                actual_val = "FAILED_TO_LOAD"
                for line in lines:
                    if re.search(r'[\u0D80-\u0DFF]', line): 
                        actual_val = line
                        break
                
                expected_val = str(ws.cell(i, 4).value or "").strip() 
                ws.cell(i, 5).value = actual_val 
                
                status = "PASS" if actual_val == expected_val else "FAIL"
                ws.cell(i, 6).value = status 

                wb.save(excel_file)

            except Exception:
                continue

        browser.close()
        
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--url", required=True)
    
    args = parser.parse_args()
    run_test(args.excel, args.url)